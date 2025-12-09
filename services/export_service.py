"""
Export Service - Raporlama ve Export servisi
PDF, Excel, CSV çıktıları
"""
from typing import List, Dict
from datetime import datetime
import os


class ExportService:
    """Export servisi - PDF, Excel, CSV çıktıları"""
    
    def __init__(self):
        pass
    
    def export_to_pdf(self, data: List[Dict], title: str, columns: List[str], filename: str = None) -> str:
        """Veriyi PDF olarak export et"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # PDF oluştur
            doc = SimpleDocTemplate(filename, pagesize=A4)
            story = []
            
            # Stil
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#667eea'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Başlık
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Tarih
            date_style = ParagraphStyle(
                'CustomDate',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1
            )
            story.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style))
            story.append(Spacer(1, 0.3*inch))
            
            # Tablo verileri
            table_data = [columns]  # Başlık satırı
            
            for row in data:
                table_row = []
                for col in columns:
                    value = row.get(col, '')
                    if isinstance(value, (int, float)):
                        if col.lower() in ['toplam', 'kdv', 'tutar', 'net', 'fiyat', 'birimfiyat']:
                            table_row.append(f"{value:.2f} ₺")
                        else:
                            table_row.append(str(value))
                    else:
                        table_row.append(str(value))
                table_data.append(table_row)
            
            # Tablo oluştur
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
            
            # PDF oluştur
            doc.build(story)
            return filename
        except Exception as e:
            raise Exception(f"PDF export hatası: {str(e)}")
    
    def export_to_excel(self, data: List[Dict], title: str, columns: List[str], filename: str = None) -> str:
        """Veriyi Excel olarak export et"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel sheet name limit
            
            # Başlık
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True, color='667EEA')
            ws.merge_cells('A1:' + get_column_letter(len(columns)) + '1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Tarih
            ws['A2'] = f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A2'].font = Font(size=10, italic=True)
            ws.merge_cells('A2:' + get_column_letter(len(columns)) + '2')
            
            # Başlık satırı
            header_row = 3
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Veri satırları
            for row_idx, row_data in enumerate(data, start=header_row + 1):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, '')
                    if isinstance(value, (int, float)):
                        if col_name.lower() in ['toplam', 'kdv', 'tutar', 'net', 'fiyat', 'birimfiyat']:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = value
                            cell.number_format = '#,##0.00" ₺"'
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            # Kolon genişliklerini ayarla
            for col_idx in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Excel export hatası: {str(e)}")
    
    def export_to_csv(self, data: List[Dict], title: str, columns: List[str], filename: str = None) -> str:
        """Veriyi CSV olarak export et"""
        try:
            import csv
            
            if not filename:
                filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                
                # Başlık
                writer.writerow([title])
                writer.writerow([f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
                writer.writerow([])
                
                # Kolon başlıkları
                writer.writerow(columns)
                
                # Veri satırları
                for row_data in data:
                    row = []
                    for col in columns:
                        value = row_data.get(col, '')
                        if isinstance(value, (int, float)):
                            if col.lower() in ['toplam', 'kdv', 'tutar', 'net', 'fiyat', 'birimfiyat']:
                                row.append(f"{value:.2f} ₺")
                            else:
                                row.append(str(value))
                        else:
                            row.append(str(value))
                    writer.writerow(row)
            
            return filename
        except Exception as e:
            raise Exception(f"CSV export hatası: {str(e)}")

